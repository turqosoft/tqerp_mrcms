import frappe

# tqerp_mrcms/api.py

import json
from frappe.utils import cint
from frappe import _
from frappe.utils.xlsxutils import make_xlsx
import frappe
from frappe.utils.xlsxutils import make_xlsx
from io import BytesIO
import math
# @frappe.whitelist(allow_guest=False)
# def submit_claim(data):
#     claim_data = json.loads(data)
#     claim = frappe.get_doc({
#         "doctype": "Claim",
#         "ip_no": claim_data.get("ip_no"),
#         "ip_name": claim_data.get("ip_name"),
#         "receipt_no": claim_data.get("receipt_no"),
#         "receipt_date": claim_data.get("receipt_date"),
#         "amount_claimed": claim_data.get("amount_claimed"),
#         "claim_type": claim_data.get("claim_type"),
#         "category": claim_data.get("category"),
#         "claim_begin_date": claim_data.get("claim_begin_date"),
#         "claim_end_date": claim_data.get("claim_end_date"),
#         "diseases": claim_data.get("diseases"),
#         "hospital": claim_data.get("hospital"),
#         "rule": claim_data.get("rule"),
#         "nominee": claim_data.get("nominee"),
#         "name_of_patient": claim_data.get("name_of_patient"),
#         "relation": claim_data.get("relation"),
#         "age_of_patient": claim_data.get("age_of_patient"),
#         "claim_templates": claim_data.get("claim_templates"),
#         "claim_checklist": claim_data.get("claim_checklist")
#     })
#     claim.insert()
#     frappe.db.commit()
#     return {"status": "success", "claim_id": claim.name}

import frappe
from frappe.utils import nowdate
from datetime import datetime

@frappe.whitelist(allow_guest=True)
def get_claim_dashboard_data():
    """
    Dashboard data for public Claims Dashboard.
    Tailored to MRCMS Claim doctype:

    Key fields:
      - claim_status (Data Entry, Registered, Processing, Sanctioned, Rejected, Returned, Paid, Closed)
      - claim_date    (Date)
      - passed_amount (Currency/Float)
    """

    # ---------- BASIC COUNTS ----------
    total_claims = frappe.db.count("Claim")

    # Status-wise counts
    status_rows = frappe.db.get_all(
        "Claim",
        fields=["claim_status", "count(*) as count"],
        group_by="claim_status"
    )
    status_counts = {r.claim_status or "Unknown": r.count for r in status_rows}

    # Make sure all expected statuses exist in the dict
    expected_statuses = [
        "Data Entry", "Registered", "Processing",
        "Sanctioned", "Rejected", "Returned",
        "Paid", "Closed"
    ]
    for s in expected_statuses:
        status_counts.setdefault(s, 0)

    # Pending = not finished (adjust as per your logic)
    pending_count = (
        status_counts.get("Data Entry", 0)
        + status_counts.get("Registered", 0)
        + status_counts.get("Processing", 0)
    )

    approved_count = status_counts.get("Sanctioned", 0)
    rejected_count = status_counts.get("Rejected", 0)
    paid_count     = status_counts.get("Paid", 0)

    # ---------- PROCESSING TIME (OVERALL AVERAGE) ----------
    # Since your Claim currently doesn't have per-stage date fields,
    # we compute a simple overall processing time:
    #   claim_date -> "completion date"
    #
    # For completed claims (Paid / Closed), use modified date as proxy.
    # For others, use today's date as proxy.
    #
    # If you later add fields like:
    #   submission_date, review_date, sanction_date, rejection_date
    # you can plug them into the SQL like we discussed earlier.

    avg_processing_days = frappe.db.sql(
        """
        SELECT AVG(
            DATEDIFF(
                CASE
                    WHEN claim_status IN ('Paid', 'Closed') THEN modified
                    ELSE %(today)s
                END,
                claim_date
            )
        ) AS avg_days
        FROM `tabClaim`
        WHERE claim_date IS NOT NULL
        """,
        {"today": nowdate()},
        as_dict=True,
    )[0].get("avg_days") or 0

    # For now, we don't have stage-specific dates in your schema.
    # We return 0 for these, and you can wire them later when you add
    # proper datetime fields for workflow stages.
    avg_sub_to_review       = 0
    avg_review_to_sanction  = 0
    avg_review_to_reject    = 0

    # ---------- FINANCIAL SUMMARY ----------
    # Using passed_amount from Claim
    finance = frappe.db.sql(
        """
        SELECT
            SUM(passed_amount) AS total_processed,
            AVG(passed_amount) AS avg_processed
        FROM `tabClaim`
        WHERE passed_amount IS NOT NULL
        """,
        as_dict=True
    )[0]

    total_processed = finance.get("total_processed") or 0
    avg_processed   = finance.get("avg_processed") or 0

    # Outstanding = sanctioned but not yet paid
    outstanding = frappe.db.sql(
        """
        SELECT SUM(passed_amount) AS outstanding
        FROM `tabClaim`
        WHERE claim_status = 'Sanctioned'
          AND passed_amount IS NOT NULL
        """,
        as_dict=True
    )[0].get("outstanding") or 0

    # ---------- CLAIM TRENDS (SUBMISSIONS OVER TIME) ----------
    # Using claim_date, grouped by year-month
    trend_rows = frappe.db.sql(
        """
        SELECT
            DATE_FORMAT(claim_date, '%%Y-%%m') AS ym,
            COUNT(*) AS count
        FROM `tabClaim`
        WHERE claim_date IS NOT NULL
        GROUP BY ym
        ORDER BY ym
        """,
        as_dict=True
    )

    trend_labels = [r.ym for r in trend_rows]
    trend_values = [r.count for r in trend_rows]

    # ---------- REIMBURSEMENT DISTRIBUTION (BUCKETS OF passed_amount) ----------
    buckets = {
        "0–5k": 0,
        "5k–10k": 0,
        "10k–25k": 0,
        "25k–50k": 0,
        "50k+": 0,
    }

    dist_rows = frappe.db.sql(
        """
        SELECT passed_amount
        FROM `tabClaim`
        WHERE passed_amount IS NOT NULL
        """,
        as_dict=True
    )

    for r in dist_rows:
        amt = float(r.passed_amount or 0)
        if amt <= 5000:
            buckets["0–5k"] += 1
        elif amt <= 10000:
            buckets["5k–10k"] += 1
        elif amt <= 25000:
            buckets["10k–25k"] += 1
        elif amt <= 50000:
            buckets["25k–50k"] += 1
        else:
            buckets["50k+"] += 1

    return {
        "summary": {
            "total_claims": total_claims,
            "pending": pending_count,
            "approved": approved_count,
            "rejected": rejected_count,
            "paid": paid_count,
            "avg_processing_days": round(avg_processing_days or 0, 2),
        },
        "status_overview": {
            "labels": list(status_counts.keys()),
            "values": list(status_counts.values()),
        },
        "financial": {
            "total_processed": float(total_processed or 0),
            "avg_processed": float(avg_processed or 0),
            "outstanding": float(outstanding or 0),
        },
        "processing_time": {
            # Currently 0 because there are no stage-specific date fields
            "sub_to_review": round(avg_sub_to_review or 0, 2),
            "review_to_sanction": round(avg_review_to_sanction or 0, 2),
            "review_to_reject": round(avg_review_to_reject or 0, 2),
        },
        "trend": {
            "labels": trend_labels,
            "values": trend_values,
        },
        "distribution": {
            "labels": list(buckets.keys()),
            "values": list(buckets.values()),
        },
    }


@frappe.whitelist()
def create_claim(ip_name=None, **kwargs):
    if not ip_name:
        frappe.throw("IP No is required.")

    # Create Claim document
    doc = frappe.new_doc("Claim")
    
    # Basic fields
    doc.ip_no = ip_name
    doc.ip_name = kwargs.get("ip_name")
    doc.receipt_no = kwargs.get("receipt_no")
    doc.receipt_date = kwargs.get("receipt_date")
    doc.amount_claimed = kwargs.get("amount_claimed")
    doc.claim_type = kwargs.get("claim_type")
    doc.category = kwargs.get("category")
    doc.claim_begin_date = kwargs.get("claim_begin_date")
    doc.claim_end_date = kwargs.get("claim_end_date")
    doc.diseases = kwargs.get("diseases")
    doc.hospital = kwargs.get("hospital")
    doc.rule = kwargs.get("rule")
    doc.nominee = kwargs.get("nominee")

    # Patient Details
    doc.name_of_patient = kwargs.get("name_of_patient")
    doc.relation = kwargs.get("relation")
    doc.age_of_patient = kwargs.get("age_of_patient")

    # Checklist
    doc.claim_templates = kwargs.get("claim_templates")
    doc.claim_checklist = kwargs.get("claim_checklist")

    doc.save(ignore_permissions=True)

    return {"status": "success", "claim_id": doc.name}


@frappe.whitelist()
def get_ip_for_claim(doctype, txt=None, searchfield=None, start=0, page_length=10, filters=None, **kwargs):
    """
    Used for ip_no link field in Claim.
    Returns [[ip_no, ip_name], ...] to show in dropdown.
    """
    filters = filters or {}

    # Get matching insured persons
    ip_list = frappe.get_all(
        "Insured Person",
        filters={"ip_no": ["like", f"%{txt}%"]},  # Search by ip_no
        fields=["name", "ip_name"],
        limit_start=start,
        limit_page_length=page_length
    )

    # Log for debugging
    frappe.log_error(message=str(ip_list), title="IP List Debug")

    # Convert to list of lists for link field dropdown
    return [[ip.name, ip.ip_name] for ip in ip_list]


@frappe.whitelist()
def get_family_members_for_claim(filters=None):
    """
    Returns family members of an Insured Person.
    filters should be JSON string: {"ip_no": "IP-0001"}
    Returns list of dicts:
    { "member_name": "...", "relation": "...", "age": ..., "nominee": "...", "hospital": "...", "rule": "..." }
    """
    if filters:
        try:
            filters = json.loads(filters)
        except Exception:
            filters = {}
    else:
        filters = {}

    ip_no = filters.get("ip_no")
    if not ip_no:
        return []

    try:
        family_doc = frappe.get_doc("Insured Person", ip_no)
    except frappe.DoesNotExistError:
        return []

    result = []

    for row in getattr(family_doc, "family_members", []):
        result.append({
            "member_name": row.member_name,
            "relation": row.relation,
            "age": getattr(row, "age_of_member", ""),
            "nominee": getattr(row, "nominee", ""),
            "hospital": getattr(row, "hospital", ""),
            "rule": getattr(row, "rule", "")
        })

    return result

import frappe

@frappe.whitelist()
def get_ip_details(ip_no):

    if not ip_no:
        return {}

    ip = frappe.get_doc("Insured Person", ip_no)

    return {
        "ip_name": ip.ip_name,
        "ip_address": ip.address,
        "employer": ip.employer,
        "phone": ip.phone,
        "dob": ip.dob,
        "gender": ip.gender,
        "dispensary": ip.dispensary,
        "family_members": ip.family_members,
        "bank_accounts": ip.bank_accounts,
        "entitlement": ip.entitlement,
        "nominee_details": ip.nominee_details
    }

@frappe.whitelist()
def get_autocomplete_list(doctype):
    """Return name list for autocomplete (no /api/resource call)."""
    try:
        records = frappe.get_all(doctype, pluck="name")
        return {"status": "success", "data": records}
    except Exception as e:
        return {"status": "error", "message": str(e)}


@frappe.whitelist()
def create_insured_person():
    import json, frappe
    try:
        data = json.loads(frappe.local.request.get_data().decode("utf-8"))
        
        ip = frappe.new_doc("Insured Person")
        ip.insurance_no = data.get("insurance_no")
        ip.ip_no = data.get("ip_no")
        ip.ip_name = data.get("ip_name")
        ip.address = data.get("address")
        ip.gender = data.get("gender")
        ip.dob = data.get("dob")
        ip.phone = data.get("phone")
        ip.local_office = data.get("local_office")
        ip.employer = data.get("employer")
        ip.nominee = data.get("nominee")

        for row in data.get("family", []):
            ip.append("family_members", row)
        for row in data.get("bank_accounts", []):
            ip.append("bank_accounts", row)
        for row in data.get("entitlement", []):
            ip.append("entitlement", row)

        ip.insert(ignore_permissions=True)
        frappe.db.commit()
        return {"status": "success", "ip_no": ip.name}

    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Create IP Error")
        return {"status": "error", "message": str(e)}



@frappe.whitelist()
def create_claim_from_ip(ip_no):

    # Load Insured Person
    ip = frappe.get_doc("Insured Person", ip_no)

    # Create new Claim
    claim = frappe.new_doc("Claim")

    # Basic field mapping
    claim.ip_no = ip.ip_no
    claim.ip_name = ip.ip_name
    claim.mobile_ph = ip.phone
    claim.address = ip.address
    claim.workplace = ip.employer

    # ---------------- NOMINEE (CHILD TABLE → SINGLE FIELD) ----------------
    # Child table name: nominee_details
    # Fields: name_of_nominee, relation_with_ip, date_of_birth, uhidabha_number, address, percentage

    if ip.nominee_details:
        # Picking the *first* nominee — you can change logic if needed
        row = ip.nominee_details[0]

        # Map only name_of_nominee to Claim.nominee
        claim.nominee = row.name_of_nominee

        # If you want to store more nominee data in claim, add more fields:
        claim.nominee_relation = row.relation_with_ip
        claim.nominee_dob = row.date_of_birth
        claim.nominee_uhid = row.uhidabha_number
        claim.nominee_address = row.address
        claim.nominee_percentage = row.percentage

    # ---------------- BANK DETAILS ----------------
    bank_row = None

    if ip.bank_accounts:
        # Prefer default account
        for row in ip.bank_accounts:
            if row.default_account:
                bank_row = row
                break

        # If no default, use first
        if not bank_row:
            bank_row = ip.bank_accounts[0]

        # Copy to Claim
        claim.bank_name = bank_row.bank
        claim.branch = bank_row.branch
        claim.ifs_code = bank_row.ifsc_code
        claim.bank_account_no = bank_row.acc_no

    # Save Claim
    claim.insert(ignore_permissions=True)
    return claim.name

@frappe.whitelist()
def get_next_claim_series():
    # This will get the next number according to the series defined in Naming Series master
    series_name = "CLM-.YYYY.-.#####"
    return frappe.model.naming.make_autoname(series_name)


@frappe.whitelist(allow_guest=False)
def check_page_permission(doc, ptype="read"):
    if "Cleark" in frappe.get_roles():
        return True
    return False

@frappe.whitelist()
def get_dashboard_stats():
    stats = {
        "total_claims": frappe.db.count("Claim"),
        "total_approved": frappe.db.count("Claim", {"claim_status": "Approved"}),
        "total_initiated": frappe.db.count("Claim", {"claim_status": "Initiated"}),
        "total_insured": frappe.db.count("Insured Person")
    }
    return stats

@frappe.whitelist()
def search_insured_persons(filters):
    """Search Insured Person records with pagination and filters.

    Expected filters (JSON string from frontend):
    {
        "page": 1,
        "page_size": 10,
        "ip_no": "123",
        "ip_name": "john",
        "phone": "9876"
    }
    """
    # Parse filters JSON from string to dict
    try:
        f = json.loads(filters)
    except Exception:
        f = {}

    page = int(f.get("page", 1)) or 1
    page_size = int(f.get("page_size", 10)) or 10
    offset = (page - 1) * page_size

    # Build conditions & args safely
    conditions = "1=1"
    args = {}

    if f.get("ip_no"):
        conditions += " AND ip_no LIKE %(ip_no)s"
        args["ip_no"] = f"%{f['ip_no']}%"

    if f.get("ip_name"):
        conditions += " AND ip_name LIKE %(ip_name)s"
        args["ip_name"] = f"%{f['ip_name']}%"

    if f.get("phone"):
        conditions += " AND phone LIKE %(phone)s"
        args["phone"] = f"%{f['phone']}%"

    # Count total (for pagination)
    # Note: frappe.db.count with filters=dict won't accept raw SQL conditions,
    # so we reuse the same WHERE manually.
    total_count = frappe.db.sql(
        f"""
        SELECT COUNT(*)
        FROM `tabInsured Person`
        WHERE {conditions}
        """,
        args,
    )[0][0]

    total_pages = (total_count // page_size) + (1 if total_count % page_size else 0)

    # Fetch paginated records
    data = frappe.db.sql(
        f"""
        SELECT
            name,
            ip_no,
            ip_name,
            phone,
            employer
        FROM `tabInsured Person`
        WHERE {conditions}
        ORDER BY ip_name
        LIMIT %(limit)s OFFSET %(offset)s
        """,
        {**args, "limit": page_size, "offset": offset},
        as_dict=True,
    )

    return {
        "list": data,
        "total_pages": total_pages,
        "total_count": total_count,
        "page": page,
        "page_size": page_size,
    }

@frappe.whitelist()
def search_claims(filters):
    """Search Claim records with pagination and filters.

    Expected filters (JSON string from frontend):
    {
        "page": 1,
        "page_size": 10,
        "from_date": "2025-01-01",
        "to_date": "2025-01-31",
        "insurance_no": "123",
        "status": "Draft"
    }
    """
    try:
        f = json.loads(filters)
    except Exception:
        f = {}

    page = int(f.get("page", 1)) or 1
    page_size = int(f.get("page_size", 10)) or 10
    offset = (page - 1) * page_size

    conditions = "1=1"
    args = {}

    # Date range filter (form_date)
    if f.get("from_date"):
        conditions += " AND claim_date >= %(from_date)s"
        args["from_date"] = f["from_date"]

    if f.get("to_date"):
        conditions += " AND claim_date <= %(to_date)s"
        args["to_date"] = f["to_date"]

    # Insurance No filter
    if f.get("insurance_no"):
        conditions += " AND ip_no LIKE %(insurance_no)s"
        args["insurance_no"] = f"%{f['insurance_no']}%"

    # Status filter
    if f.get("status"):
        conditions += " AND claim_status = %(status)s"
        args["status"] = f["status"]

    # Total count for pagination
    total_count = frappe.db.sql(
        f"""
        SELECT COUNT(*)
        FROM `tabClaim`
        WHERE {conditions}
        """,
        args,
    )[0][0]

    total_pages = (total_count // page_size) + (1 if total_count % page_size else 0)

    # Fetch page of claims
    data = frappe.db.sql(
        f"""
        SELECT
            name,
            ip_no,
            ip_name,
            name_of_patient,
            age_of_patient,
            relation,
            amount_claimed,
            claim_status,
            claim_date
        FROM `tabClaim`
        WHERE {conditions}
        ORDER BY claim_date DESC, creation DESC
        LIMIT %(limit)s OFFSET %(offset)s
        """,
        {**args, "limit": page_size, "offset": offset},
        as_dict=True,
    )

    return {
        "list": data,
        "total_pages": total_pages,
        "total_count": total_count,
        "page": page,
        "page_size": page_size,
    }
@frappe.whitelist()
def save_claim():
    import json
    data = frappe.local.form_dict  # or json.loads(frappe.local.request.get_data())

    claim_name = data.get("claim_name")

    if claim_name:
        doc = frappe.get_doc("Claim", claim_name)  # EDIT existing
    else:
        doc = frappe.new_doc("Claim")  # CREATE new

    # -------------------------
    # Update main/master fields
    # -------------------------
    main_fields = [
        "ip_no", "phone", "address", "ip_name", "book_no", "employer",
        "name_of_patient", "age_of_patient", "relation", "from_date", "to_date",
        "type", "hospital", "in_patient_no", "diseases", "amount_claimed",
        "bank_name", "branch", "bank_account_no", "ifs_code", "claim_status",
        "claim_date", "form_no", "passed_amount", "passed_amount_words",
        "voucher_no", "remarks"
    ]
    for field in main_fields:
        if field in data:
            doc.set(field, data[field])

    # -------------------------
    # Update child tables
    # -------------------------
    if "bill_details" in data:
        doc.set("bill_details", data["bill_details"])
    if "contacted_ip" in data:
        doc.set("contacted_ip", data["contacted_ip"])

    # -------------------------
    # Save the document
    # -------------------------
    doc.save()
    frappe.db.commit()  # commit changes
    return {"claim_name": doc.name}


@frappe.whitelist()
def get_ip_details(ip_no: str):
    """Fetch Insured Person + family details for pre-filling the claim form."""
    if not ip_no:
        return None

    # Adjust the DocType and fieldnames as per your actual schema
    ip_doc = frappe.get_doc("Insured Person", {"ip_no": ip_no})

    return frappe._dict(
        ip_no=ip_doc.ip_no,
        phone=ip_doc.phone,
        ip_name=ip_doc.ip_name,
        address=getattr(ip_doc, "address", ""),
        employer=getattr(ip_doc, "employer", ""),
        bank_name=getattr(ip_doc, "bank_name", ""),
        branch=getattr(ip_doc, "branch", ""),
        bank_ac_no=getattr(ip_doc, "bank_ac_no", ""),
        ifsc=getattr(ip_doc, "ifsc", ""),
        dispensary=getattr(ip_doc, "dispensary", ""),
        family_members=ip_doc.get("family_members") or []
    )


@frappe.whitelist()
def get_claim_by_form_no():
    claim_no = frappe.form_dict.get("claim_no")  # this is the doc.name
    if not claim_no:
        return {"error": "claim_no missing"}

    try:
        doc = frappe.get_doc("Claim", claim_no)

        return {
            "claim": doc.as_dict(),
            "bill_details": [row.as_dict() for row in doc.get("bill_details") or []],
            "contacted_ip": [row.as_dict() for row in doc.get("ip_communication") or []]
        }

    except Exception:
        frappe.log_error(frappe.get_traceback(), "get_claim_by_form_no")
        return {"error": "Error fetching claim details, check logs"}



import frappe, json

@frappe.whitelist(allow_guest=False)
def update_claim():
    """
    Update an existing Claim document from JSON payload.
    Expects child tables as arrays of dicts:
    - bill_details: [{"bill_date":..., "bill_no":..., "bill_amount":...}]
    - contacted_ip: [{"contacted_ip_date":..., "contacted_ip_by":..., "contacted_ip_remarks":...}]
    """
    try:
        # -------------------------
        # Get payload
        # -------------------------
        data = frappe.request.get_json() or frappe.local.form_dict or {}
        frappe.log_error(json.dumps(data, default=str), "Update Claim Payload")

        # Use form_no or name as document identifier
        doc_name = data.get("name") or data.get("form_no")
        if not doc_name:
            return {"error": "Document identifier (form_no) is required for update"}

        # -------------------------
        # Fetch existing claim
        # -------------------------
        if not frappe.db.exists("Claim", doc_name):
            return {"error": f"Claim '{doc_name}' does not exist"}

        doc = frappe.get_doc("Claim", doc_name)

        # -------------------------
        # Correct mapping (HTML -> DocType)
        # -------------------------
        field_map = {
            "insurance_no": "ip_no",
            "mobile": "phone",
            "ip_address": "address",
            "ip_name": "ip_name",
            "esi_book_no": "book_no",
            "employer_name": "employer",
            "patient_name": "name_of_patient",   # ✅ Make sure fieldname matches DocType
            "patient_age": "age_of_patient",
            # "relationship": "relation",
            "treat_from": "from_date",
            "treat_to": "to_date",
            "treatment_type": "type",
            "ip_treatment_details": "hospital",
            "in_patient_no": "in_patient_no",
            "diagnosis": "diseases",
            "claim_amount": "amount_claimed",
            "bank_name": "bank_name",
            "branch": "branch",
            "bank_ac_no": "bank_account_no",
            "ifsc": "ifs_code",
            "claim_status": "claim_status",
            "claim_date": "claim_date",
            "form_no": "name",
            "passed_amount": "passed_amount",
            "passed_amount_words": "passed_amount_words",
            "voucher_no": "voucher_no",
            "remarks": "remarks"
        }

        # -------------------------
        # Update main fields safely
        # -------------------------
        for html_field, dt_field in field_map.items():
            if html_field in data:
                val = data.get(html_field)
                if isinstance(val, str):
                    val = val[:140]  # truncate long strings
                if hasattr(doc, dt_field):
                    doc.set(dt_field, val or "")
                    frappe.log_error(f"Setting {dt_field} = {val}", "Update Claim Debug")

        # -------------------------
        # Update Bill Details child table
        # -------------------------
        if isinstance(data.get("bill_details"), list):
            doc.set("bill_details", [])
            for row in data["bill_details"]:
                doc.append("bill_details", {
                    "bill_date": row.get("bill_date") or None,
                    "bill_no": str(row.get("bill_no") or "")[:140],
                    "bill_amount": float(str(row.get("bill_amount") or 0).replace(",", ""))
                })

        # -------------------------
        # Update Contacted IP child table
        # -------------------------
        if isinstance(data.get("contacted_ip"), list):
            doc.set("ip_communication", [])
            for row in data["contacted_ip"]:
                if not any([row.get("contacted_ip_date"), row.get("contacted_ip_by"), row.get("contacted_ip_remarks")]):
                    continue
                doc.append("ip_communication", {
                    "date": row.get("contacted_ip_date") or None,
                    "name1": (row.get("contacted_ip_by") or "")[:140],
                    "remarks": (row.get("contacted_ip_remarks") or "")[:140]
                })

        # -------------------------
        # Save and commit
        # -------------------------
        doc.save(ignore_permissions=True)
        frappe.db.commit()

        return {"name": doc.name}

    except Exception:
        frappe.log_error(frappe.get_traceback(), "Update Claim Failed")
        return {"error": "Failed to update claim. Check server logs."}

@frappe.whitelist()
def get_ip_details_list(doctype, txt, searchfield, start, page_len, filters):
    return frappe.db.sql("""
        SELECT
            ip_no,
            CONCAT(MAX(ip_name), ' - ', MAX(phone), ' - ', MAX(dispensary))
        FROM `tabInsured Person`
        WHERE
            ip_no LIKE %(txt)s
            OR ip_name LIKE %(txt)s
            OR phone LIKE %(txt)s
        GROUP BY ip_no
        ORDER BY MAX(ip_name) ASC
        LIMIT %(start)s, %(page_len)s
    """, {
        "txt": f"%{txt}%",
        "start": start,
        "page_len": page_len
    })




import frappe
from datetime import datetime, date

@frappe.whitelist()
def get_family_members_for_dropdown(ip_no):
    """Return list of family member names for a given Insured Person"""
    if not ip_no:
        return []

    try:
        ip_doc = frappe.get_doc("Insured Person", ip_no)
        return [f.member_name for f in getattr(ip_doc, "family_members", [])]
    except frappe.DoesNotExistError:
        return []

@frappe.whitelist()
def get_family_member_details(ip_no, member_name):
    """Return relation and age_of_patient for selected family member"""
    if not ip_no or not member_name:
        return {}

    try:
        ip_doc = frappe.get_doc("Insured Person", ip_no)
        for f in getattr(ip_doc, "family_members", []):
            if f.member_name == member_name:
                # Calculate age from dob
                age = None
                if f.dob:
                    dob = f.dob
                    if isinstance(dob, str):
                        dob = datetime.strptime(dob, "%Y-%m-%d").date()
                    today = date.today()
                    age = today.year - dob.year - ((today.month, today.day) < (dob.month, dob.day))

                return {
                    "relation": f.relation or "",
                    "age_of_patient": age
                }
        return {}
    except frappe.DoesNotExistError:
        return {}




import frappe
import json

@frappe.whitelist()
def create_claim_proceeding_for_multiple(claims_data):
    """
    Create ONE Claim Proceedings document with all selected claims in the child table
    """
    if not claims_data:
        frappe.throw("No claims selected.")

    # Parse JSON string if needed
    if isinstance(claims_data, str):
        claims_data = json.loads(claims_data)

    # Get logged-in user's office (adjust fieldname on User if different)
    user_office = frappe.db.get_value("User", frappe.session.user, "office")
    emp_full_name = frappe.db.get_value("User", frappe.session.user, "full_name")

    # Create parent document
    cp = frappe.get_doc({
        "doctype": "Claim Proceedings",
        "naming_series": "CP-.YYYY.-",  # <-- set on parent
        "claim_proceedings": []         # child table fieldname
    })
    # Set office on the parent if available
    if user_office:
        # adjust "office" to your actual fieldname on Claim Proceedings
        cp.office = user_office
    # Set logged in user as employee on the parent if available
    cp.employee = frappe.session.user
    if emp_full_name:
        cp.employee_name = emp_full_name

    # Append each selected claim to the child table
    for claim in claims_data:
        cp.append("claim_proceedings", {  # must be child table fieldname
        "claim_no": claim.get("claim_no", ""),
            "name1": claim.get("ip_name", ""),
            "ip_no": claim.get("ip_no", ""),
            "claim_date": claim.get("claim_date", ""),
            "phone": claim.get("phone", ""),
            "ifsc": claim.get("ifs_code", ""),
            "bank_account_no": claim.get("bank_account_no", ""),
            "passed_amount": claim.get("passed_amount", 0)
        })

    cp.insert(ignore_permissions=True)
    return {"name": cp.name}
# your_app/api.py

import frappe

@frappe.whitelist()
def get_claim_dates_for_print(claim_name):
    """
    Returns earliest and latest claim_date for all claim_no in claim_proceedings child table
    """
    doc = frappe.get_doc("Claim", claim_name)
    dates = []

    for row in doc.claim_proceedings:
        if row.claim_no:
            try:
                claim_doc = frappe.get_doc("Claim", row.claim_no)
                if claim_doc.claim_date:
                    dates.append(claim_doc.claim_date)
            except frappe.DoesNotExistError:
                continue

    if dates:
        earliest = min(dates)
        latest = max(dates)
        return {"earliest": earliest, "latest": latest}
    else:
        return {"earliest": None, "latest": None}


import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from frappe.utils import get_site_path
import frappe

# @frappe.whitelist()
# def download_claim_details_excel(docname):
#     """
#     Download Excel for Claim Proceedings with fields:
#     claim_no, name1, ip_no, phone, ifsc, bank_account_no, passed_amount, claim_date
#     """
#     doc = frappe.get_doc("Claim Proceedings", docname)  # Parent DocType

#     filename = f"{docname}_claim_details.xlsx"
#     filepath = get_site_path("public", "files", filename)

#     wb = openpyxl.Workbook()
#     ws = wb.active
#     ws.title = "Claim Details"

#     # Styles
#     grey_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
#     bold_font = Font(bold=True)

#     # Row 1: Company heading
#     ws.merge_cells('A1:H1')
#     ws['A1'] = "MRCMS"
#     ws['A1'].font = Font(bold=True, size=14)
#     ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
#     ws['A1'].fill = grey_fill

#     # Row 2: Report Heading
#     ws.merge_cells('A2:H2')
#     ws['A2'] = "Claim Proceedings Details"
#     ws['A2'].font = Font(bold=True, size=12)
#     ws['A2'].alignment = Alignment(horizontal="center", vertical="center")
#     ws['A2'].fill = grey_fill

#     ws.append([])  # Blank row

#     # Optional: Parent-level info (if any)
#     ws.append(["Document Name", doc.name])
#     ws.append(["Generated On", frappe.utils.now_datetime().strftime("%d-%m-%Y %H:%M")])
#     ws.append([])

#     # Headers
#     headers = ["Claim No", "Name", "IP No", "Phone", "IFSC", "Bank Account No", "Passed Amount", "Claim Date"]
#     ws.append(headers)

#     # Style header row
#     for cell in ws[ws.max_row]:
#         cell.fill = grey_fill
#         cell.font = bold_font
#         cell.alignment = Alignment(horizontal="center")

#     # Populate data
#     total_passed_amount = 0.0
#     for row in doc.claim_proceedings:
#         ws.append([
#             row.claim_no or "",
#             row.name1 or "",
#             row.ip_no or "",
#             row.phone or "",
#             row.ifsc or "",
#             row.bank_account_no or "",
#             float(row.passed_amount or 0),
#             row.claim_date.strftime("%d-%m-%Y") if row.claim_date else ""
#         ])
#         total_passed_amount += float(row.passed_amount or 0)

#     ws.append([])

#     # Total row
#     ws.append(["", "", "", "", "", "TOTAL", total_passed_amount, ""])
#     for cell in ws[ws.max_row]:
#         cell.font = bold_font
#         if cell.column_letter == "G":
#             cell.alignment = Alignment(horizontal="right")

#     # Column widths
#     column_widths = {
#         "A": 15, "B": 30, "C": 15, "D": 15,
#         "E": 15, "F": 20, "G": 15, "H": 15
#     }
#     for col_letter, width in column_widths.items():
#         ws.column_dimensions[col_letter].width = width

#     wb.save(filepath)
#     return f"/files/{filename}"

import csv
from frappe.utils import get_site_path, now_datetime

@frappe.whitelist()
def download_claim_details_excel(docname):
 
    doc = frappe.get_doc("Claim Proceedings", docname)
 
    filename = f"{docname}_claim_details.xlsx"
    filepath = get_site_path("public", "files", filename)
 
    # Remove old file
    if os.path.exists(filepath):
        os.remove(filepath)
 
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Claim Details"
 
    # Header row only
    headers = ["Claim No", "Name", "IP No", "Phone", "IFSC", "ACCOUNT NO.", "Passed Amount", "Claim Date"]
    ws.append(headers)
 
    # Style header
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
 
    # Populate rows
    total_passed_amount = 0.0
    for row in doc.claim_proceedings:
        ws.append([
            row.claim_no or "",
            row.name1 or "",
            row.ip_no or "",
            row.phone or "",
            row.ifsc or "",
            row.bank_account_no or "",
            float(row.passed_amount or 0),
            row.claim_date.strftime("%d-%m-%Y") if row.claim_date else ""
        ])
        total_passed_amount += float(row.passed_amount or 0)
 
    # Total row
    ws.append(["", "", "", "", "", "TOTAL", total_passed_amount, ""])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
 
    # Column widths
    widths = {"A": 15, "B": 30, "C": 15, "D": 15, "E": 15, "F": 20, "G": 15, "H": 15}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
 
    wb.save(filepath)
    return f"/files/{filename}"

import pandas as pd
import frappe
from frappe.utils import get_site_path
 
@frappe.whitelist()
def download_claim_details_csv(docname):
    """
    Download CSV for Claim Proceedings with fields:
    claim_no, name1, ip_no, phone, ifsc, bank_account_no, passed_amount, claim_date
    """
 
    doc = frappe.get_doc("Claim Proceedings", docname)  # Parent DocType
 
    filename = f"{docname}_claim_details.csv"
    filepath = get_site_path("public", "files", filename)
 
    data = []
 
    total_passed_amount = 0.0
 
    # Header row
    headers = [
        "Claim No", "Name", "IP No", "Phone",
        "IFSC", "Bank Account No", "Passed Amount", "Claim Date"
    ]
 
    # Populate rows
    for row in doc.claim_proceedings:
        passed_amount = float(row.passed_amount or 0)
 
        data.append([
            row.claim_no or "",
            row.name1 or "",
            row.ip_no or "",
            row.phone or "",
            row.ifsc or "",
            row.bank_account_no or "",
            passed_amount,
            row.claim_date.strftime("%d-%m-%Y") if row.claim_date else ""
        ])
 
        total_passed_amount += passed_amount
 
    # Add total row
    data.append(["", "", "", "", "", "TOTAL", total_passed_amount, ""])
 
    # Create DataFrame
    df = pd.DataFrame(data, columns=headers)
 
    # Save CSV
    df.to_csv(filepath, index=False)
 
    return f"/files/{filename}"


def update_claim_status_on_submit(doc, method=None):
    """
    On submit of Claim Proceedings:
    Set each linked Claim's claim_status = "Proceedings"
    """
    if doc.docstatus == 1:  # Submitted
        for row in doc.claim_proceedings:
            if row.claim_no:
                frappe.db.set_value("Claim", row.claim_no, "claim_status", "Proceedings")
    frappe.db.commit()




# def update_claim_status_on_paid(doc, method=None):
#     """
#     On update of Claim Proceedings:
#     If proceedings_status is 'Paid', set each linked Claim's claim_status = "Paid"
#     """

#     frappe.errprint(f"update_claim_status_on_paid triggered for doc: {doc.name}")
#     frappe.errprint(f"proceedings_status: {doc.proceedings_status}, docstatus: {doc.docstatus}")

#     # Skip if document is cancelled
#     if doc.docstatus == 2:
#         frappe.errprint(f"Document {doc.name} is cancelled. Skipping Paid update.")
#         return

#     if doc.docstatus == 1 and doc.proceedings_status == "Paid":
#         for row in doc.claim_proceedings:
#             frappe.errprint(f"Processing claim_no: {row.claim_no}")
#             if row.claim_no:
#                 claim_doc = frappe.get_doc("Claim", row.claim_no)
#                 frappe.errprint(f"Current claim_status: {claim_doc.claim_status}")

#                 if claim_doc.claim_status != "Paid":
#                     frappe.db.set_value("Claim", row.claim_no, "claim_status", "Paid")
#                     frappe.errprint(f"Updated claim_status to Paid for: {row.claim_no}")
#                 else:
#                     frappe.errprint(f"Already Paid: {row.claim_no}")


def update_claim_status_on_cancel(doc, method=None):
    """
    On cancel of Claim Proceedings:
    Set each linked Claim's claim_status = 'Sanctioned'
    """
    frappe.errprint(f"update_claim_status_on_cancel triggered for doc: {doc.name}")

    for row in doc.claim_proceedings:
        if row.claim_no:
            claim_doc = frappe.get_doc("Claim", row.claim_no)
            frappe.errprint(f"Current claim_status: {claim_doc.claim_status}")
            
            frappe.db.set_value("Claim", row.claim_no, "claim_status", "Sanctioned")
            frappe.errprint(f"Updated claim_status to 'Sanctioned' for: {row.claim_no}")

import frappe
import pandas as pd
from frappe.utils import getdate

@frappe.whitelist()
def process_payment_file_org(docname, file_url):
    """
    Reads uploaded Excel/CSV with columns:
    MOBILE, ACCOUNT NO., AMOUNT, CREDIT DATE, CREDIT STATUS, UTR Number

    Matches each row in child table (doc.claims) using:
      - bank_account_no
      - passed_amount

    Updates:
      - credit_date
      - credit_status
      - utr_number
      - credit_amount
    """

    doc = frappe.get_doc("Claim Proceedings", docname)

    # Resolve file path from /files/ URL
    # e.g. /files/payment.xlsx -> sites/site/public/files/payment.xlsx
    
    filename = file_url.split("/files/")[-1]
    file_path = frappe.get_site_path("public", "files", filename)

    ext = file_path.split(".")[-1].lower()

    if ext in ("xlsx", "xls"):
        df = pd.read_excel(file_path)
    elif ext == "csv":
        df = pd.read_csv(file_path)
    else:
        frappe.throw("Unsupported file type. Please upload CSV or Excel.")

    # Normalise headers: lower, trim, replace spaces & dots
    # "ACCOUNT NO."   -> "account_no"
    # "CREDIT DATE"   -> "credit_date"
    # "UTR Number"    -> "utr_number"
    df.columns = [
        c.strip().lower().replace(" ", "_").replace(".", "")
        for c in df.columns
    ]

    # We expect these normalised column names
    required_cols = ["account_no", "amount", "credit_date", "credit_status", "utr_number"]
    for col in required_cols:
        if col not in df.columns:
            frappe.throw(f"Missing required column in uploaded file: {col}")

    rows = df.to_dict("records")

    updated = 0
    unmatched = []

    # Assume child table is doc.claims; change if your fieldname is different
    for row in doc.claim_proceedings:
        matched = False
        for rec in rows:
            # Compare on account_no + amount == passed_amount
            bank_acc_file = str(rec.get("account_no") or "").strip()
            bank_acc_row  = str(row.bank_account_no or "").strip()

            amt_file = float(rec.get("amount") or 0)
            amt_row  = float(row.passed_amount or 0)

            if bank_acc_row == bank_acc_file and amt_row == amt_file:
                # Match found: update payment fields
                row.credit_amount = amt_file
                row.credit_status = rec.get("credit_status")
                row.utr_number    = rec.get("utr_number")

                credit_date_val = rec.get("credit_date")
                if credit_date_val:
                    row.credit_date = getdate(credit_date_val)

                frappe.db.set_value("Claim", row.claim_no, "claim_status", "Paid")

                updated += 1
                matched = True
                break

        if not matched:
            # Track unmatched rows (for info)
            unmatched.append({
                "bank_account_no": row.bank_account_no,
                "passed_amount": row.passed_amount
            })

    doc.save(ignore_permissions=True)
    frappe.db.commit()

    return {
        "updated": updated,
        "unmatched": unmatched
    }

@frappe.whitelist()
def process_payment_file(docname, file_url):
    import os
    import pandas as pd
    from frappe.utils import getdate

    doc = frappe.get_doc("Claim Proceedings", docname)

    if doc.docstatus != 1:
        frappe.throw("Upload allowed only after submission.")

    # ---------------------
    # Resolve File Path
    # ---------------------
    filename = file_url.split("/files/")[-1]
    file_path = frappe.get_site_path("public", "files", filename)

    if not os.path.exists(file_path):
        frappe.throw(f"File not found: {file_path}")

    ext = file_path.split(".")[-1].lower()

    if ext in ("xlsx", "xls"):
        df = pd.read_excel(file_path)
    else:
        df = pd.read_csv(file_path)

    df.columns = [
        c.strip().lower().replace(" ", "_").replace(".", "")
        for c in df.columns
    ]

    required_cols = ["account_no", "amount", "credit_date", "credit_status", "utr_number"]
    for col in required_cols:
        if col not in df.columns:
            frappe.throw(f"Missing column: {col}")

    rows = df.to_dict("records")

    updated = 0
    unmatched = []

    for row in doc.claim_proceedings:
        matched = False
        for rec in rows:
            if (str(row.bank_account_no).strip() ==
                str(rec.get("account_no") or "").strip()
                and float(row.passed_amount) ==
                    float(rec.get("amount") or 0)):

                row.credit_amount = rec.get("amount")
                row.credit_status = rec.get("credit_status")
                row.utr_number = rec.get("utr_number")
                row.credit_date = getdate(rec.get("credit_date"))
                row._highlight = "green"

                # -----------------------
                # Update Claims to To Paid
                # -----------------------
                frappe.db.set_value("Claim", row.claim_no, "claim_status", "Paid")

                updated += 1
                matched = True
                break

        if not matched:
            unmatched.append({
                "bank_account_no": row.bank_account_no,
                "passed_amount": row.passed_amount
            })

    # -----------------------
    # Generate Mismatch Excel
    # -----------------------
    mismatch_file_url = None
    if unmatched:
        mismatch_df = pd.DataFrame(unmatched)
        mismatch_filename = f"mismatch_{docname}.xlsx"
        mismatch_path = frappe.get_site_path("public", "files", mismatch_filename)
        mismatch_df.to_excel(mismatch_path, index=False)
        mismatch_file_url = f"/files/{mismatch_filename}"

    # -----------------------
    # Log Upload History
    # -----------------------
    # doc.append("upload_history", {
    #     "file_name": filename,
    #     "file_url": file_url,
    #     "uploaded_on": frappe.utils.now(),
    #     "processed_by": frappe.session.user,
    #     "updated_count": updated,
    #     "unmatched_count": len(unmatched),
    # })

    # -----------------------
    # Auto Move To Paid
    # -----------------------
    if all([r.credit_status and str(r.credit_status).lower() == "delivered"
            for r in doc.claim_proceedings]):
        doc.proceedings_status = "Paid"

    doc.save(ignore_permissions=True)
    frappe.db.commit()

    return {
        "updated": updated,
        "unmatched": unmatched,
        "unmatched_count": len(unmatched),
        "mismatch_file_url": mismatch_file_url
    }

@frappe.whitelist()
def number_to_words_indian(num):
    """Convert a number to Indian words (Camel Case / Title Case)"""
   
    try:
        num = int(math.floor(float(num)))
    except:
        return ""
 
    if num == 0:
        return "Zero"
 
    ones = ["", "One", "Two", "Three", "Four", "Five", "Six", "Seven", "Eight", "Nine",
            "Ten", "Eleven", "Twelve", "Thirteen", "Fourteen", "Fifteen", "Sixteen",
            "Seventeen", "Eighteen", "Nineteen"]
    tens = ["", "", "Twenty", "Thirty", "Forty", "Fifty", "Sixty", "Seventy", "Eighty", "Ninety"]
 
    def two_digit_to_words(n):
        if n < 20:
            return ones[n]
        t = n // 10
        o = n % 10
        return tens[t] + (" " + ones[o] if o else "")
 
    def three_digit_to_words(n):
        h = n // 100
        rest = n % 100
        return (ones[h] + " Hundred " if h else "") + (two_digit_to_words(rest) if rest else "")
 
    crore = num // 10000000
    num %= 10000000
    lakh = num // 100000
    num %= 100000
    thousand = num // 1000
    hundreds_and_rest = num % 1000
 
    parts = []
    if crore:
        parts.append(three_digit_to_words(crore) + " Crore")
    if lakh:
        parts.append(three_digit_to_words(lakh) + " Lakh")
    if thousand:
        parts.append(three_digit_to_words(thousand) + " Thousand")
    if hundreds_and_rest:
        parts.append(three_digit_to_words(hundreds_and_rest))
 
    # Join all parts and capitalize the first letter of each word
    result = " ".join(parts).strip()
    return result + " Only"


# -----claim bundle management creation-------
@frappe.whitelist()
def create_claim_bundle_management(claims_data):
    import json
 
    if not claims_data:
        frappe.throw("⚠️ No claims selected. Please select at least one claim.")
 
    if isinstance(claims_data, str):
        claims_data = json.loads(claims_data)
 
    duplicate_claims = []
    valid_claims = []
 
    for claim in claims_data:
        claim_no = claim.get("claim_no")
       
        # Get all Claim Bundle Details for this claim
        bundle_details = frappe.get_all(
            "Claim Bundle Details",
            filters={"claim_no": claim_no},
            fields=["parent"]
        )
 
        # Flag to check if claim is submitted in any bundle
        is_submitted = False
 
        for bd in bundle_details:
            cbm_status = frappe.db.get_value("Claim Bundle Management", bd.parent, "docstatus")
            if cbm_status == 1:  # Submitted
                is_submitted = True
                break
 
        if is_submitted:
            duplicate_claims.append(claim_no)
        else:
            valid_claims.append(claim)
 
    # Show message for duplicate claims
    if duplicate_claims:
        duplicate_str = ", ".join([f"<b>{d}</b>" for d in duplicate_claims])
        frappe.msgprint(
            f"⚠️ The following claim(s) are already submitted in a Claim Bundle: {duplicate_str}",
            title="Duplicate Claims",
            indicator="red"
        )
 
    # Stop if no valid claims
    if not valid_claims:
        return
 
    # Create new Claim Bundle with only valid claims
    user_office = frappe.db.get_value("User", frappe.session.user, "office")
 
    cbm = frappe.get_doc({
        "doctype": "Claim Bundle Management",
        "office": user_office,
        "details": []
    })
 
    for claim in valid_claims:
        cbm.append("details", {
            "claim_no": claim.get("claim_no") or "",
            "claim_date": claim.get("claim_date", ""),
            "ip_no": claim.get("ip_no", ""),
            "ip_name": claim.get("ip_name", ""),
            "phone": claim.get("phone", ""),
            "name_of_patient": claim.get("name_of_patient", ""),
            "dispensary": claim.get("dispensary", ""),
            "claim_status": claim.get("claim_status"),
            "amount_claimed": claim.get("amount_claimed", 0),
            "passed_amount": claim.get("passed_amount", 0),
            "ifs_code": claim.get("ifs_code", 0),
            "bank_account_no": claim.get("bank_account_no", 0),
            "bank_name": claim.get("bank_name", 0)
        })
 
    cbm.insert(ignore_permissions=True)
 
    frappe.msgprint(
        f"✅ Claim Bundle <b>{cbm.name}</b> created successfully!",
        title="Success",
        indicator="green"
    )
 
    return {
        "name": cbm.name,
        "redirect_to": f"/app/claim-bundle-management/{cbm.name}"
    } 
 
# -----claim Sanction list------
@frappe.whitelist()
def create_claim_sanction_list(claims_data):
    """
    claims_data: list of claim details, each with claim_bundle_no (from parent CBM)
    """
    import json
 
    if not claims_data:
        frappe.throw("⚠️ No claim bundle rows selected.")
 
    if isinstance(claims_data, str):
        claims_data = json.loads(claims_data)
 
    duplicate_claims = []
    valid_claims = []
 
    # --- VALIDATION: Skip claims whose Claim Bundle is already in a submitted CSL ---
    for row in claims_data:
        claim_no = row.get("claim_no")
        claim_bundle_no = row.get("claim_bundle_no")
 
        if not claim_no:
            frappe.throw("Claim No missing for some row.")
 
        # Check if this claim_bundle_no is already in a submitted Claim Sanction List
        existing_rows = frappe.get_all(
            "Claim Sanction Details",  
            filters={"claim_bundle_no": claim_bundle_no, "claim_no": claim_no},
            fields=["parent"]
        )
 
        is_submitted = False
        for er in existing_rows:
            status = frappe.db.get_value("Claim Sanction List", er.parent, "docstatus")
            if status == 1:  # Submitted
                is_submitted = True
                break
 
        if is_submitted:
            duplicate_claims.append(claim_bundle_no)  # <-- store bundle number instead of claim no
        else:
            valid_claims.append(row)
 
    # Show message for duplicate claims (displaying bundle numbers)
    if duplicate_claims:
        duplicate_str = ", ".join([f"<b>{d}</b>" for d in duplicate_claims])
        frappe.msgprint(
            f"⚠️ The following Claim Bundle(s) are already submitted in a Claim Sanction List: {duplicate_str}",
            title="Duplicate Claim Bundles",
            indicator="red"
        )
 
    # Stop if no valid claims
    if not valid_claims:
        return
 
    # Get logged-in user's office
    user_office = frappe.db.get_value("User", frappe.session.user, "office")
 
    # Create parent Claim Sanction List
    csl = frappe.get_doc({
        "doctype": "Claim Sanction List",
        "office": user_office,
        "details": []  # child table
    })
 
    # Append only valid claims
    for row in valid_claims:
        csl.append("details", {
            "claim_bundle_no": row.get("claim_bundle_no"),
            "claim_no": row.get("claim_no"),
            "claim_date": row.get("claim_date", ""),
            "ip_no": row.get("ip_no", ""),
            "ip_name": row.get("ip_name", ""),
            "phone": row.get("phone", ""),
            "name_of_patient": row.get("name_of_patient", ""),
            "dispensary": row.get("dispensary", ""),
            "claim_status": row.get("claim_status", ""),
            "amount_claimed": row.get("amount_claimed", 0),
            "passed_amount": row.get("passed_amount", 0),
            "ifs_code": row.get("ifs_code", 0),
            "bank_account_no": row.get("bank_account_no", 0),
            "bank_name": row.get("bank_name", 0)
        })
 
    csl.insert(ignore_permissions=True)
 
    frappe.msgprint(
        f"✅ Claim Sanction List <b>{csl.name}</b> created successfully!",
        title="Success",
        indicator="green"
    )
 
    # Redirect to the new CSL
    return {
        "name": csl.name,
        "redirect_to": f"/app/claim-sanction-list/{csl.name}"
    }
 
# -----claim Payment List-------
@frappe.whitelist()
def create_claim_payment_list(payments_data):
    """
    Create ONE Claim Payment List document
    with multiple rows in Payment Details child table
    and skip already submitted claims
    """
    import json
 
    if not payments_data:
        frappe.throw("⚠️ No claim sanction rows selected.")
 
    # Convert JSON string to list
    if isinstance(payments_data, str):
        payments_data = json.loads(payments_data)
 
    duplicate_claims = []
    valid_claims = []
 
    # --- VALIDATION: Skip claims whose Claim Sanction is already in a submitted CPL ---
    for row in payments_data:
        claim_sanction_no = row.get("claim_sanction_no")
 
        if not claim_sanction_no:
            frappe.throw("Claim Sanction No missing for some row.")
 
        # Check if this claim_sanction_no is already in any submitted CPL
        existing_rows = frappe.get_all(
            "Claim Payment Details",  # child table of Claim Payment List
            filters={"claim_sanction_no": claim_sanction_no},
            fields=["parent"]
        )
 
        is_submitted = False
        for er in existing_rows:
            status = frappe.db.get_value("Claim Payment List", er.parent, "docstatus")
            if status == 1:  # Submitted
                is_submitted = True
                break
 
        if is_submitted:
            duplicate_claims.append(claim_sanction_no)
        else:
            valid_claims.append(row)
 
    # Show message for duplicate claims
    if duplicate_claims:
        duplicate_str = ", ".join([f"<b>{d}</b>" for d in duplicate_claims])
        frappe.msgprint(
            f"⚠️ The following Claim Sanction(s) are already submitted in a Claim Payment List: {duplicate_str}",
            title="Duplicate Claims",
            indicator="red"
        )
 
    # Stop if no valid claims
    if not valid_claims:
        return
 
    # Logged-in user's office
    user_office = frappe.db.get_value("User", frappe.session.user, "office")
 
    # Create parent document
    cpl = frappe.get_doc({
        "doctype": "Claim Payment List",
        "office": user_office,
        "details": []  
    })
 
    # Add only valid claims
    for row in valid_claims:
        cpl.append("details", {
            "claim_no": row.get("claim_no") or "",
            "claim_date": row.get("claim_date", ""),
            "ip_no": row.get("ip_no", ""),
            "ip_name": row.get("ip_name", ""),
            "phone": row.get("phone", ""),
            "name_of_patient": row.get("name_of_patient", ""),
            "dispensary": row.get("dispensary", ""),
            "claim_status": row.get("claim_status", ""),
            "amount_claimed": row.get("amount_claimed", 0),
            "passed_amount": row.get("passed_amount", 0),
            "ifs_code":row.get("ifs_code", 0),
            "bank_account_no": row.get("bank_account_no", 0),
            "bank_name": row.get("bank_name", 0),
            "claim_sanction_no": row.get("claim_sanction_no", 0)
        })
 
    # Save document
    cpl.insert(ignore_permissions=True)
 
    frappe.msgprint(
        f"✅ Claim Payment List <b>{cpl.name}</b> created successfully!",
        title="Success",
        indicator="green"
    )
 
    return {
        "name": cpl.name,
        "redirect_to": f"/app/claim-payment-list/{cpl.name}"
    }


def validate_fund_availability(doc):
    if not doc.fund_manager:
        return
 
    total_allocated = flt(doc.total_allocated or 0)
 
    fm_doc = frappe.get_doc("Fund Manager", doc.fund_manager)
 
    total_fixed = sum(flt(row.fixed or 0) for row in fm_doc.details)
 
    total_previous_allocated = frappe.db.sql("""
        SELECT COALESCE(SUM(total_allocated), 0)
        FROM `tabClaim Payment List`
        WHERE fund_manager = %s
          AND docstatus = 1
          AND name != %s
    """, (doc.fund_manager, doc.name))[0][0]
 
    fund_available = total_fixed - total_previous_allocated
 
    if total_allocated > fund_available:
        frappe.throw(
            f"Total Allocated ({total_allocated}) exceeds current Available Fund ({fund_available})."
        )
   
def validate(doc, method):
    if doc.doctype not in ("Claim Payment List", "Claim Proceedings"):
        return
 
    validate_fund_availability(doc)

#download payment list as excel
import frappe
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from frappe.utils import get_site_path
import os
 
@frappe.whitelist()
def download_payment_details_excel(docname):
 
     
    doc = frappe.get_doc("Claim Payment List", docname)
 
    # Prepare file path
    filename = f"{docname}_payment_details.xlsx"
    filepath = get_site_path("public", "files", filename)
 
 
    # Remove old file
    if os.path.exists(filepath):
        os.remove(filepath)
 
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Payment Details"
 
    # Header row only
    headers = ["Claim No", "Name", "IP No", "Phone", "IFSC", "Bank Account No", "Passed Amount", "Claim Date"]
    ws.append(headers)
 
    # Style header
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
 
    # Populate rows
    total_passed_amount = 0.0
    for row in doc.details:
        ws.append([
            row.claim_no or "",
            row.ip_name or "",
            row.ip_no or "",
            row.phone or "",
            row.ifs_code or "",
            row.bank_account_no or "",
            float(row.passed_amount or 0),
            row.claim_date.strftime("%d-%m-%Y") if row.claim_date else ""
        ])
        total_passed_amount += float(row.passed_amount or 0)
 
    # Total row
    ws.append(["", "", "", "", "", "TOTAL", total_passed_amount, ""])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
 
    # Column widths
    widths = {"A": 15, "B": 30, "C": 15, "D": 15, "E": 15, "F": 20, "G": 15, "H": 15}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
 
    wb.save(filepath)
    return f"/files/{filename}"
 
 
 
import pandas as pd
import frappe
from frappe.utils import get_site_path
 
@frappe.whitelist()
def download_payment_details_csv(docname):
    """
    Download CSV for Claim Proceedings with fields:
    claim_no, name1, ip_no, phone, ifsc, bank_account_no, passed_amount, claim_date
    """
 
    doc = frappe.get_doc("Claim Payment List", docname)
 
    # Prepare file path
    filename = f"{docname}_payment_details.csv"
    filepath = get_site_path("public", "files", filename)
 
    data = []
 
    total_passed_amount = 0.0
 
    # Header row
    headers = [
        "Claim No", "Name", "IP No", "Phone",
        "IFSC", "Bank Account No", "Passed Amount", "Claim Date"
    ]
 
    # Populate rows
    for row in doc.details:
        passed_amount = float(row.passed_amount or 0)
 
        data.append([
           row.claim_no or "",
            row.ip_name or "",
            row.ip_no or "",
            row.phone or "",
            row.ifs_code or "",
            row.bank_account_no or "",
            float(row.passed_amount or 0),
            row.claim_date.strftime("%d-%m-%Y") if row.claim_date else ""
        ])
 
        total_passed_amount += passed_amount
 
    # Add total row
    data.append(["", "", "", "", "", "TOTAL", total_passed_amount, ""])
 
    # Create DataFrame
    df = pd.DataFrame(data, columns=headers)
 
    # Save CSV
    df.to_csv(filepath, index=False)
 
    return f"/files/{filename}"

@frappe.whitelist()
def process_payment_file_paymentlist(docname, file_url):
    import os
    import pandas as pd
    from frappe.utils import getdate
 
    doc = frappe.get_doc("Claim Payment List", docname)
 
    if doc.docstatus != 1:
        frappe.throw("Upload allowed only after submission.")
        # ---------------------
    # Validate File URL
    # ---------------------
    if not file_url:
        frappe.throw("No file selected. Please upload a payment file.")
 
    filename = file_url.split("/files/")[-1]
 
    # Build paths
    public_path = frappe.get_site_path("public", "files", filename)
    private_path = frappe.get_site_path("private", "files", filename)
 
    # Detect actual path
    if os.path.exists(public_path):
        file_path = public_path
    elif os.path.exists(private_path):
        file_path = private_path
    else:
        frappe.throw(f"""
            File not found:<br><br>
            <b>Public:</b> {public_path}<br>
            <b>Private:</b> {private_path}<br><br>
            Ensure the file is uploaded correctly.
        """)
 
    # ---------------------
    # Read Excel/CSV File
    # ---------------------
    ext = file_path.split(".")[-1].lower()
 
    try:
        if ext in ("xlsx", "xls"):
            df = pd.read_excel(file_path)
        else:
            df = pd.read_csv(file_path)
    except Exception as e:
        frappe.throw(f"Failed to read the file: {e}")
 
 
    df.columns = [
        c.strip().lower().replace(" ", "_").replace(".", "")
        for c in df.columns
    ]
 
    required_cols = ["account_no", "amount", "credit_date", "credit_status", "utr_number"]
    for col in required_cols:
        if col not in df.columns:
            frappe.throw(f"Missing column: {col}")
 
    rows = df.to_dict("records")
 
    updated = 0
    unmatched = []
 
    for row in doc.details:
        matched = False
        for rec in rows:
            if (str(row.bank_account_no).strip() ==
                str(rec.get("account_no") or "").strip()
                and float(row.passed_amount) ==
                    float(rec.get("amount") or 0)):
 
                row.credit_amount = rec.get("amount")
                row.credit_status = rec.get("credit_status")
                row.utr_number = rec.get("utr_number")
                row.credit_date = getdate(rec.get("credit_date"))
                row._highlight = "green"
 
                # -----------------------
                # Update Claims to To Paid
                # -----------------------
                frappe.db.set_value("Claim", row.claim_no, "claim_status", "Paid")
 
                updated += 1
                matched = True
                break
 
        if not matched:
            unmatched.append({
                "bank_account_no": row.bank_account_no,
                "passed_amount": row.passed_amount
            })
 
    # -----------------------
    # Generate Mismatch Excel
    # -----------------------
    mismatch_file_url = None
    if unmatched:
        mismatch_df = pd.DataFrame(unmatched)
        mismatch_filename = f"mismatch_{docname}.xlsx"
        mismatch_path = frappe.get_site_path("public", "files", mismatch_filename)
        mismatch_df.to_excel(mismatch_path, index=False)
        mismatch_file_url = f"/files/{mismatch_filename}"
 
   
    # -----------------------
    # Auto Move To Paid
    # -----------------------
    if all([r.credit_status and str(r.credit_status).lower() == "delivered"
            for r in doc.details]):
        doc.payment_status = "Paid"
 
    doc.save(ignore_permissions=True)
    frappe.db.commit()
 
    return {
        "updated": updated,
        "unmatched": unmatched,
        "unmatched_count": len(unmatched),
        "mismatch_file_url": mismatch_file_url
    }

# Funding Details updation
from frappe.utils import flt
import frappe
 
@frappe.whitelist()
def get_fixed_fund_for_office(office):
    """Return sum of fixed fund for the latest submitted Fund Manager for the office"""
    if not office:
        frappe.throw("Office is required")
 
    fm_list = frappe.get_all(
        "Fund Manager",
        filters={"office": office, "docstatus": 1},
        order_by="`tabFund Manager`.modified desc",
        limit_page_length=1,
        fields=["name"]
    )
 
    if not fm_list:
        return {"fixed": 0}
 
    fm_doc = frappe.get_doc("Fund Manager", fm_list[0].name)
    fixed_total = sum([flt(row.fixed or 0) for row in fm_doc.details])
 
    return {"fixed": fixed_total}
 
 
# @frappe.whitelist()
# def allocate_fund_on_submit(payment_list_name):
#     """Update Fund Manager and Payment List balance on submit"""
#     payment_doc = frappe.get_doc("Claim Payment List", payment_list_name)
#     office = payment_doc.office
#     total_allocated = flt(payment_doc.payment_total or 0)  # Use payment_total
 
#     if total_allocated <= 0:
#         frappe.throw("Payment Total must be greater than 0")
 
#     # Get latest submitted Fund Manager
#     fm_list = frappe.get_all(
#         "Fund Manager",
#         filters={"office": office, "docstatus": 1},
#         order_by="`tabFund Manager`.modified desc",
#         limit_page_length=1,
#         fields=["name"]
#     )
 
#     if not fm_list:
#         frappe.throw("No submitted Fund Manager found for this office")
 
#     fm_doc = frappe.get_doc("Fund Manager", fm_list[0].name)
 
#     # Allocate amount row-wise
#     remaining = total_allocated
#     for row in fm_doc.details:
#         # calculate available in row
#         available_in_row = flt(row.fixed or 0) - flt(row.allocated or 0)
#         allocate = min(available_in_row, remaining)
#         row.allocated = flt(row.allocated or 0) + allocate
#         row.paid = flt(row.paid or 0) + allocate
#         row.allocatable = flt(row.fixed or 0) - flt(row.allocated or 0)
#         remaining -= allocate
 
#         # Force update in DB directly for safety
#         frappe.db.set_value("Fund Manager Details", row.name, {
#             "allocated": row.allocated,
#             "allocatable": row.allocatable,
#             "paid": row.paid
#         })
 
#         if remaining <= 0:
#             break
 
#     if remaining > 0:
#         frappe.throw("Payment Total exceeds available fixed fund in Fund Manager")
 
#     fm_doc.reload()  # refresh parent doc
#     fm_doc.save(ignore_permissions=True)
#     frappe.db.commit()
 
#     # Update Payment List
#     payment_doc.total_allocated = total_allocated
#     payment_doc.balance = flt(payment_doc.available or 0) - total_allocated
#     payment_doc.save(ignore_permissions=True)
#     frappe.db.commit()
 
#     return True

@frappe.whitelist()
def get_claim_category_by_amount(passed_amount):
    amount = float(passed_amount)
 
    categories = frappe.get_all(
        "Claim Category",
        fields=["name", "min_amount", "max_amount"],
        order_by="min_amount asc"  
    )
 
    for c in categories:
        min_val = float(c.min_amount or 0)
        max_val = float(c.max_amount or 0)
 
        if min_val <= amount <= max_val:
            return c.name
 
    return None

# -------------------------------------------------
# Allocate fund on CPL submit (SAFE SYNC)
# -------------------------------------------------
@frappe.whitelist()
def allocate_fund_on_submit(docname, doctype=None):
    """Freeze CPL balance and safely sync Fund Manager"""
 
    if not doctype:
        doctype = "Claim Payment List"
 
    doc = frappe.get_doc(doctype, docname)
 
    validate_fund_availability(doc)
 
    fm_name = doc.fund_manager
    total_allocated = flt(doc.total_allocated or 0)
 
    if not fm_name:
        frappe.throw("Please select a Fund Manager before submitting.")
 
    if total_allocated <= 0:
        frappe.throw("Total Allocated must be greater than 0")
 
    # -----------------------------
    # FUND MANAGER – SOURCE VALUES
    # -----------------------------
    fm_doc = frappe.get_doc("Fund Manager", fm_name)
 
    total_fixed = sum(flt(row.fixed or 0) for row in fm_doc.details)
 
    total_previous_allocated = frappe.db.sql("""
        SELECT COALESCE(SUM(total_allocated), 0)
        FROM `tabClaim Payment List`
        WHERE fund_manager = %s
          AND docstatus = 1
          AND name != %s
    """, (fm_name, doc.name))[0][0]
 
    fund_available = total_fixed - total_previous_allocated
 
    if total_allocated > fund_available:
        frappe.throw(
            f"Total Allocated ({total_allocated}) exceeds Available Fund ({fund_available})."
        )
 
    # -----------------------------
    # FREEZE BALANCE IN CPL
    # -----------------------------
    doc.balance = fund_available - total_allocated
    # doc.db_set("balance", fund_available - total_allocated)
 
 
    if hasattr(doc, "payment_status"):
        doc.payment_status = "Paid"
 
    if hasattr(doc, "proceedings_status"):
        doc.proceedings_status = "Paid"
 
    # -----------------------------
    # SAFE FUND MANAGER SYNC
    # -----------------------------
    remaining = total_allocated
 
    for row in fm_doc.details:
        # ✅ ONLY MATCHING OFFICE
        if row.office != doc.office:
            continue
 
        fixed = flt(row.fixed or 0)
        allocated = flt(row.allocated or 0)
 
        available_in_row = fixed - allocated
        if available_in_row <= 0:
            frappe.throw(
                f"No available fund for office {doc.office}"
            )
 
        consume = min(available_in_row, remaining)
 
        row.allocated = allocated + consume
        row.paid = flt(row.paid or 0) + consume
        row.allocatable = fixed - row.allocated
 
        remaining -= consume
        break  # STOP after correct office
 
    fm_doc.save(ignore_permissions=True)
 
    return True

# -------------------------------------------------
# Get fund details (FOR UI)
# -------------------------------------------------
@frappe.whitelist()
def get_fund_details(fund_manager, office=None):
    """
    Returns available fund for a specific Fund Manager and office.
    If office is not provided, returns total fund across all offices.
    """
    if not fund_manager:
        return {}
 
    fm_doc = frappe.get_doc("Fund Manager", fund_manager)
 
    # Filter by office if provided
    if office:
        details = [d for d in fm_doc.details if d.office == office]
    else:
        details = fm_doc.details
 
    total_fixed = sum(flt(d.fixed or 0) for d in details)
 
    # Sum allocated amounts for submitted Claim Payment List for this fund manager and office
    if office:
        total_allocated = frappe.db.sql("""
            SELECT COALESCE(SUM(total_allocated), 0)
            FROM `tabClaim Payment List`
            WHERE fund_manager = %s
              AND office = %s
              AND docstatus = 1
        """, (fund_manager, office))[0][0]
    else:
        total_allocated = frappe.db.sql("""
            SELECT COALESCE(SUM(total_allocated), 0)
            FROM `tabClaim Payment List`
            WHERE fund_manager = %s
              AND docstatus = 1
        """, fund_manager)[0][0]
 
    available = total_fixed - total_allocated
 
    return {
        "available": available,
        "fund_date": fm_doc.get("date"),
        "approval_note": fm_doc.approval_note or ""
    }

# -------------------------------------------------
# Reverse fund on CPL cancel
# -------------------------------------------------
@frappe.whitelist()
def reverse_fund_on_cancel(payment_list_name=None, doctype=None, name=None, **kwargs):
    """
    Reverse Fund Manager values when Claim Payment List or Claim Proceedings is cancelled
    """
 
    # Determine correct document name
    docname = payment_list_name or name
    if not docname:
        return True
 
    # Detect correct DocType
    if doctype:
        doc = frappe.get_doc(doctype, docname)
    else:
        doc = frappe.get_doc("Claim Payment List", docname)
 
    refund_amount = flt(doc.total_allocated or 0)
    if refund_amount <= 0:
        return True
 
    if not doc.fund_manager or not doc.office:
        return True
 
    fm_doc = frappe.get_doc("Fund Manager", doc.fund_manager)
 
    reversed_done = False
 
    for row in fm_doc.details:
        if row.office != doc.office:
            continue
 
        allocated = flt(row.allocated or 0)
        if allocated <= 0:
            continue
 
        refund = min(allocated, refund_amount)
 
        row.db_set("allocated", allocated - refund)
        row.db_set("paid", flt(row.paid or 0) - refund)
        row.db_set(
            "allocatable",
            flt(row.fixed or 0) - (allocated - refund)
        )
 
        reversed_done = True
        break  #  office-wise safety
 
    if not reversed_done:
        frappe.log_error(
            f"No allocation found to reverse for {doc.doctype} {doc.name}",
            "Fund Reverse Warning"
        )
 
    fm_doc.save(ignore_permissions=True)
    fm_doc.reload()
 
    return True
 
# -------------------------------------------------
# Get available Fund Managers for office
# -------------------------------------------------
@frappe.whitelist()
def get_available_fund_managers(doctype=None, txt=None, searchfield=None,
                                start=0, page_len=20, filters=None, **kwargs):
 
    import json
    if filters and isinstance(filters, str):
        filters = json.loads(filters)
 
    office = filters.get("office") if filters else None
    if not office:
        return []
 
    fm_details = frappe.get_all(
        "Fund Manager Details",
        filters={"office": office},
        fields=["parent"]
    )
 
    fm_names = list(set(d.parent for d in fm_details))
 
    result = []
    for fm_name in fm_names:
        fm_doc = frappe.get_doc("Fund Manager", fm_name)
        total_allocatable = sum(
            flt(row.fixed or 0) - flt(row.allocated or 0)
            for row in fm_doc.details
        )
        if total_allocatable > 0:
            result.append([fm_doc.name, fm_doc.get("date") or ""])
 
    return result